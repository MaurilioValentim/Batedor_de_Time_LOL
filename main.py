import openpyxl
from time import sleep
import random
import itertools

# Malks,Ajala,Ale,Lucas,Leo,Bernardo,Arthur,Mika,Dudu,Guguin
## ==================================================================

pontos_vencedor = 1
pontos_perdedor = 1
pontos_dos_n_jogadores = 0.5
disparidade = 2

## ==================================================================


planilha_pontuacao = openpyxl.load_workbook('Pontuacao.xlsx')
pagina_dados = planilha_pontuacao['Plan1']

nome = []
pontos = []

for i,linha in enumerate(pagina_dados.iter_rows(min_row=2)):

    nome.append(linha[0].value)
    pontos.append(linha[1].value)

valores = dict(zip(nome, pontos))

opcao = input("Bater time (1) ou adicionar vencedores (2): ")

opcao = int(opcao)

if opcao == 1:
    print("Você escolheu bater time.")
    
    dados = []
    pessoas = []

    pessoas = input("Digite o nome das 10 pessoas que iram participar (separadas por vírgula): ")

    # Separar os nomes digitados em uma lista
    pessoas = [p.strip() for p in pessoas.split(',')]

    # Verificar se os nomes estão no dicionário e pegar os pontos correspondentes
    for pessoa in pessoas:
        if pessoa in valores:
            dados.append((pessoa, valores[pessoa]))  
        else:
            print(f"{pessoa} não está na lista de pontos.")

   # Calcular a soma total de pontos
    soma_total = sum(ponto for _, ponto in dados)

    # Definir o limite máximo de disparidade
    limite_disparidade = disparidade  # Diferença máxima permitida entre as somas das equipes

    # Tentar encontrar duas equipes de 5 pessoas cada com a menor diferença possível
    nomes = [pessoa for pessoa, _ in dados]  # Lista dos nomes
    random.shuffle(nomes)  # Embaralhar a lista de nomes para tornar o processo aleatório

    # Variáveis para armazenar a melhor combinação encontrada
    melhor_disparidade = float('inf')
    melhor_equipe1 = []
    melhor_equipe2 = []

    # Gerar todas as combinações possíveis de 5 jogadores
    for combinacao in itertools.combinations(nomes, 5):
        # Calcular a soma de pontos da combinação atual
        pontos_combinacao1 = sum(valores[nome] for nome in combinacao)
        equipe2 = [nome for nome in nomes if nome not in combinacao]
        pontos_combinacao2 = sum(valores[nome] for nome in equipe2)

        # Calcular a disparidade entre as duas equipes
        disparidade = abs(pontos_combinacao1 - pontos_combinacao2)

        # Verificar se a disparidade é menor ou igual ao limite definido
        if disparidade <= limite_disparidade and disparidade < melhor_disparidade:
            melhor_disparidade = disparidade
            melhor_equipe1 = list(combinacao)
            melhor_equipe2 = equipe2

    # Exibir os resultados
    if melhor_equipe1 and melhor_equipe2:
        pontos_equipe1 = sum(valores[nome] for nome in melhor_equipe1)
        pontos_equipe2 = sum(valores[nome] for nome in melhor_equipe2)
        
        print(f"Equipe 1: {melhor_equipe1}, Soma: {pontos_equipe1}")
        print(f"Equipe 2: {melhor_equipe2}, Soma: {pontos_equipe2}")
        print(f"Disparidade: {abs(pontos_equipe1 - pontos_equipe2)}")
    else:
        print("Não foi possível encontrar duas equipes com disparidade aceitável.")

elif opcao == 2:
    print("Você escolheu adicionar vencedores.")

    ganhadores = input("Digite o nome dos 5 ganhadores (separadas por vírgula): ")

    perdedores = input("Digite o nome dos 5 perdedores (separadas por vírgula): ")

    for i,jogadores in enumerate(pagina_dados.iter_rows(min_row=2)):
        if jogadores[0].value in ganhadores:
            jogadores[1].value += pontos_vencedor
        elif jogadores[0].value in perdedores:
            jogadores[1].value -= pontos_perdedor 
        else: 
            jogadores[1].value -= pontos_dos_n_jogadores         

    planilha_pontuacao.save('Pontuacao.xlsx')    
                
else:
    print("Opção inválida.")


